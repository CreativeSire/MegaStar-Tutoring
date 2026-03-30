from __future__ import annotations

from datetime import date, datetime, timedelta
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from typing import Dict, Iterable
from urllib.parse import parse_qs

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = BASE_DIR / "Brashan Chemistry Staff Invoice.xlsx"
HTML_PATH = BASE_DIR / "invoice_builder.html"
HOST = "127.0.0.1"
PORT = 8765


def parse_date(value: str | None, fallback: date | None = None) -> date:
    if value:
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            pass
    return fallback or date.today()


def month_label(value: date) -> str:
    return value.strftime("%B").upper()


def previous_month_label(value: date) -> str:
    previous_month = value.replace(day=1) - timedelta(days=1)
    return month_label(previous_month)


def safe_text(value: str | None) -> str:
    return (value or "").strip()


def safe_number(value: str | None):
    text = safe_text(value)
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return text
    return int(number) if number.is_integer() else number


def pair_rows(*lists: Iterable[str]):
    materialized = [list(values) for values in lists]
    longest = max((len(values) for values in materialized), default=0)
    for index in range(longest):
        yield tuple(values[index] if index < len(values) else "" for values in materialized)


def build_filename(invoice_date: date, invoice_no: str | None = None) -> str:
    _ = invoice_no
    base = f"BRASHAN INVOICE {invoice_date.strftime('%B %Y').upper()}"
    return base.replace("/", "-").replace("\\", "-").replace(":", "-") + ".xlsx"


def set_payment_details(ws, payment: Dict[str, str]) -> None:
    ws["B41"] = payment.get("account_name", "")
    ws["B42"] = payment.get("account_number", "")
    ws["B43"] = payment.get("swift_code", "")
    ws["B44"] = payment.get("bank_name", "")
    ws["B45"] = payment.get("sort_code", "")
    ws["B46"] = payment.get("location", "")


def set_line_items(ws, descriptions: list[str], quantities: list[str], prices: list[str]) -> None:
    rows = range(16, 31)
    for row, description, quantity, unit_price in zip(rows, descriptions, quantities, prices):
        ws[f"B{row}"] = description
        ws[f"D{row}"] = safe_number(quantity)
        ws[f"E{row}"] = safe_number(unit_price)
        ws[f"F{row}"] = f'=IF(D{row}="",ROUND(1*E{row},2),ROUND(D{row}*E{row},2))'

    for row in range(16 + len(descriptions), 31):
        ws[f"B{row}"] = None
        ws[f"D{row}"] = None
        ws[f"E{row}"] = None
        ws[f"F{row}"] = f'=IF(D{row}="",ROUND(1*E{row},2),ROUND(D{row}*E{row},2))'


def set_report_rows(ws, report_dates: list[str], report_students: list[str]) -> None:
    rows = range(51, 86)
    for row, report_date, report_student in zip(rows, report_dates, report_students):
        ws[f"A{row}"] = safe_text(report_date)
        ws[f"C{row}"] = safe_text(report_student)

    for row in range(51 + len(report_dates), 86):
        ws[f"A{row}"] = None
        ws[f"C{row}"] = None


def populate_workbook(form: Dict[str, list[str]]) -> tuple[BytesIO, str]:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Missing template: {TEMPLATE_PATH}")

    invoice_date = parse_date(form.get("invoice_date", [None])[0], date.today())
    invoice_no = safe_text(form.get("invoice_no", [""])[0])

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Invoice"]

    ws["E8"] = invoice_no
    ws["F8"] = invoice_date.strftime("%d/%m/%Y")
    ws["E11"] = safe_text(form.get("terms", ["Due upon receipt"])[0]) or "Due upon receipt"
    ws["A48"] = f"REPORT OF PREVIOUS MONTH ({previous_month_label(invoice_date)})"

    payment = {
        "account_name": safe_text(form.get("account_name", [""])[0]),
        "account_number": safe_text(form.get("account_number", [""])[0]),
        "swift_code": safe_text(form.get("swift_code", [""])[0]),
        "bank_name": safe_text(form.get("bank_name", [""])[0]),
        "sort_code": safe_text(form.get("sort_code", [""])[0]),
        "location": safe_text(form.get("location", [""])[0]),
    }
    set_payment_details(ws, payment)

    descriptions = form.get("line_desc", [])
    quantities = form.get("line_qty", [])
    unit_prices = form.get("line_unit_price", [])
    set_line_items(ws, descriptions, quantities, unit_prices)

    report_dates = form.get("report_date", [])
    report_students = form.get("report_student", [])
    set_report_rows(ws, report_dates, report_students)

    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, build_filename(invoice_date, invoice_no)


def html_page() -> bytes:
    return HTML_PATH.read_bytes()


class InvoiceHandler(BaseHTTPRequestHandler):
    def _send(self, status: int, content_type: str, body: bytes, extra_headers: dict[str, str] | None = None) -> None:
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        if extra_headers:
            for key, value in extra_headers.items():
                self.send_header(key, value)
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        if self.path in {"/", "/index.html"}:
            self._send(200, "text/html; charset=utf-8", html_page())
            return
        if self.path == "/health":
            self._send(200, "text/plain; charset=utf-8", b"ok")
            return
        self._send(404, "text/plain; charset=utf-8", b"Not found")

    def do_POST(self):
        if self.path != "/generate":
            self._send(404, "text/plain; charset=utf-8", b"Not found")
            return

        length = int(self.headers.get("Content-Length", "0"))
        payload = self.rfile.read(length).decode("utf-8")
        form = parse_qs(payload, keep_blank_values=True)

        try:
            workbook, filename = populate_workbook(form)
        except Exception as exc:
            body = f"Failed to generate workbook: {exc}".encode("utf-8")
            self._send(500, "text/plain; charset=utf-8", body)
            return

        self.send_response(200)
        self.send_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(workbook.getbuffer())))
        self.end_headers()
        self.wfile.write(workbook.getvalue())

    def log_message(self, format, *args):
        return


def main() -> None:
    server = ThreadingHTTPServer((HOST, PORT), InvoiceHandler)
    print(f"Invoice builder running at http://{HOST}:{PORT}")
    print("Open that address in your browser, fill the form, then click Generate Excel.")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopping.")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
